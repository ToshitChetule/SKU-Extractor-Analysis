import pandas as pd
import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

def run_variant_analysis(input_file, output_dir="outputs"):
    """
    Runs variant attribute analysis for uploaded Excel file.
    Saves the consolidated output and chart to an Excel file in output_dir.
    """
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, f"processed_{os.path.basename(input_file)}")

    # Columns that should NOT be analyzed as attributes
    non_attribute_columns = ['SKU ID', 'SKU Descriptions', 'Revenue']

    df = pd.read_excel(input_file)
    df.columns = df.columns.str.strip()

    # Determine attribute columns
    attribute_columns = [col for col in df.columns if col not in non_attribute_columns]

    consolidated = []

    for attribute in attribute_columns:
        df_attr = df[df[attribute].notna()]
        if df_attr.empty:
            continue

        summary = (
            df_attr.groupby(attribute)
            .agg({"Revenue": "sum", "SKU ID": "nunique"})
            .rename(columns={
                "Revenue": "Total Annual Revenue",
                "SKU ID": "# of Unique Customers"
            })
            .reset_index()
        )

        total_rev = summary["Total Annual Revenue"].sum()
        summary["Percentage of Total Revenue"] = (summary["Total Annual Revenue"] / total_rev).round(4)
        summary["Attribute"] = attribute
        summary = summary.rename(columns={attribute: "Values"})
        summary = summary[
            ["Attribute", "Values", "Total Annual Revenue", "Percentage of Total Revenue", "# of Unique Customers"]
        ].sort_values(["Attribute", "Total Annual Revenue"], ascending=[True, False])
        consolidated.append(summary)

    if not consolidated:
        raise ValueError("No valid attribute columns found in file.")

    final_table = pd.concat(consolidated, ignore_index=True)

    # Write to Excel (percent in decimal form for Excel % formatting)
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        final_table.to_excel(writer, index=False, sheet_name="Data_Analysis")

    # Autofit columns/rows and apply number formats
    _autofit_and_format_excel(output_file, sheet_name="Data_Analysis")

    # Add charts to a new sheet (aggregated by Attribute)
    #add_bar_chart_to_excel(output_file, sheet_name="Data_Analysis", chart_sheet_name="Bar Chart 1")
    add_advanced_horizontal_percent_chart(output_file, sheet_name="Data_Analysis", chart_sheet_name="Bar Graph 1")

    return output_file  # ✅ return path for download

def _autofit_and_format_excel(filepath, sheet_name="Data_Analysis"):
    """
    Adjust column widths/heights to fit content, and apply number formats.
    """
    wb = load_workbook(filepath)
    if sheet_name not in wb.sheetnames:
        wb.save(filepath)
        return
    ws = wb[sheet_name]

    headers = [cell.value for cell in ws[1]]

    # Autofit column widths based on max length of cell text in each column
    for col_cells in ws.columns:
        col_idx = col_cells[0].column
        col_letter = get_column_letter(col_idx)
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        ws.column_dimensions[col_letter].width = max(8, max_length + 2)

    # Autofit row heights by line breaks
    for row in ws.iter_rows():
        max_lines = max((cell.value.count("\n") + 1) if isinstance(cell.value, str) else 1 for cell in row)
        if max_lines > 1:
            ws.row_dimensions[row[0].row].height = max_lines * 15

    # Apply wrap_text & vertical alignment to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Apply number formats
    try:
        rev_col_idx = headers.index("Total Annual Revenue") + 1
    except ValueError:
        rev_col_idx = None
    try:
        pct_col_idx = headers.index("Percentage of Total Revenue") + 1
    except ValueError:
        pct_col_idx = None

    if rev_col_idx:
        for row in ws.iter_rows(min_row=2, min_col=rev_col_idx, max_col=rev_col_idx):
            for cell in row:
                cell.number_format = '$#,##0.00'
    if pct_col_idx:
        for row in ws.iter_rows(min_row=2, min_col=pct_col_idx, max_col=pct_col_idx):
            for cell in row:
                cell.number_format = '0.0%'

    wb.save(filepath)

# def add_bar_chart_to_excel(filepath, sheet_name="Data_Analysis", chart_sheet_name="Bar Chart"):
#     """
#     Adds a bar chart with one bar per Attribute (sum of revenues over values, in millions),
#     showing only the total revenue value data label on each bar (formatted as $Xm).
#     """
#     wb = load_workbook(filepath)
#     ws = wb[sheet_name]

#     # Collect data and aggregate
#     data = []
#     headers = [cell.value for cell in ws[1]]
#     attr_idx = headers.index("Attribute")
#     rev_idx = headers.index("Total Annual Revenue")
#     for row in ws.iter_rows(min_row=2, values_only=True):
#         data.append({
#             "Attribute": row[attr_idx],
#             "Total Annual Revenue": (row[rev_idx] or 0) / 1_000_000  # Scale to millions
#         })
#     df = pd.DataFrame(data)
#     chart_df = (
#         df.groupby("Attribute", as_index=False)
#           .agg({"Total Annual Revenue": "sum"})
#     )
#     data_rows = len(chart_df)

#     # Remove old chart/data sheets
#     for s in ["ChartData", chart_sheet_name]:
#         if s in wb.sheetnames:
#             del wb[s]
#     # Write chart data
#     cd_ws = wb.create_sheet("ChartData")
#     cd_ws.append(["Attribute", "Total Revenue (Millions)"])
#     for row in chart_df.itertuples(index=False):
#         cd_ws.append(list(row))

#     # Format chart data cells to show $Xm style
#     for row in cd_ws.iter_rows(min_row=2, min_col=2, max_col=2, max_row=data_rows+1):
#         for cell in row:
#             cell.number_format = '$0.0,"M"'  # Data label will inherit this format

#     # Chart setup
#     bar_chart = BarChart()
#     bar_chart.type = "col"
#     bar_chart.title = "Total Revenue by Attribute"
#     bar_chart.y_axis.title = "Total Revenue (Millions)"
#     bar_chart.x_axis.title = "Attribute"

#     # Remove grid lines
#     bar_chart.y_axis.majorGridlines = None
#     bar_chart.x_axis.majorGridlines = None

#     # Only add revenue (1 series)
#     rev_data = Reference(cd_ws, min_col=2, min_row=1, max_row=data_rows+1)
#     categories = Reference(cd_ws, min_col=1, min_row=2, max_row=data_rows+1)
#     bar_chart.add_data(rev_data, titles_from_data=True)
#     bar_chart.set_categories(categories)

#     # Show value only (the number) on top of each bar
#     bar_chart.dataLabels = DataLabelList()
#     bar_chart.dataLabels.showVal = True
#     bar_chart.dataLabels.showSerName = False
#     bar_chart.dataLabels.showCatName = False
#     bar_chart.dataLabels.separator = "\n"  # Optional: Separate by newline, for clarity

#     # Expand chart size, optional
#     bar_chart.width = 22
#     bar_chart.height = 14

#     # Add space between chart border and bars
#     bar_chart.gapWidth = 250
#     bar_chart.overlap = 0

#     # Explicitly show axes and tick marks
#     bar_chart.x_axis.majorTickMark = "in"
#     bar_chart.x_axis.minorTickMark = "in"
#     bar_chart.x_axis.visible = True
#     bar_chart.y_axis.majorTickMark = "in"
#     bar_chart.y_axis.minorTickMark = "in"
#     bar_chart.y_axis.visible = True

#     # Show legend
#     bar_chart.legend = None  # Only one series, so legend not needed

#     # New chart sheet
#     chart_ws = wb.create_sheet(chart_sheet_name)
#     chart_ws.add_chart(bar_chart, "B2")

#     # Hide the ChartData sheet
#     cd_ws.sheet_state = "hidden"

#     wb.save(filepath)

def add_advanced_horizontal_percent_chart(filepath, sheet_name="Data_Analysis", chart_sheet_name="Bar Graph 1"):
    wb = load_workbook(filepath)
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    val_col = headers.index("Values")
    pct_col = headers.index("Percentage of Total Revenue")
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append({
            "Value": row[val_col],
            "Percentage": row[pct_col]
        })
    df = pd.DataFrame(data)
    # Remove rows where percentage is zero
    df = df[df["Percentage"] > 0]
    # Sort by percentage descending and reset index
    df = df.sort_values("Percentage", ascending=False).reset_index(drop=True)
    # Reverse DataFrame so highest percentage is at the bottom (top of chart)
    df = df.iloc[::-1].reset_index(drop=True)
    # Write chart data to sheet
    pct_ws_name = "PercentChartData"
    for s in [pct_ws_name, chart_sheet_name]:
        if s in wb.sheetnames:
            del wb[s]
    pct_ws = wb.create_sheet(pct_ws_name)
    pct_ws.append(["Attribute Value", "Percentage of Total Revenue"])
    for row in df.itertuples(index=False):
        pct_ws.append([row.Value, row.Percentage])
    # Format % on percentage column
    nrows = len(df)
    for row in pct_ws.iter_rows(min_row=2, min_col=2, max_col=2, max_row=nrows+1):
        for c in row:
            c.number_format = '0.0%'
    # Build simple horizontal bar chart
    bar = BarChart()
    bar.type = "bar"
    bar.title = "% Revenue by Attribute Value"
    bar.x_axis.title = "Values"
    bar.x_axis.number_format = '0.0%'
    bar.y_axis.title = "Percentage of Total Revenue"
    # Explicitly show axes and tick marks
    bar.x_axis.majorTickMark = "in"
    bar.x_axis.minorTickMark = "in"
    bar.x_axis.visible = True
    bar.y_axis.majorTickMark = "in"
    bar.y_axis.minorTickMark = "in"
    bar.y_axis.visible = True
    # Add legend
    #bar.legend = True
    # Remove grid lines
    bar.x_axis.majorGridlines = None
    bar.y_axis.majorGridlines = None
    # Add data
    data_ref = Reference(pct_ws, min_col=2, min_row=1, max_row=nrows+1)
    cats = Reference(pct_ws, min_col=1, min_row=2, max_row=nrows+1)
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats)
    # Show percentage value at end of each bar
    bar.dataLabels = DataLabelList()
    bar.dataLabels.showVal = True
    bar.dataLabels.showLegendKey = False
    bar.dataLabels.showCatName = False
    bar.dataLabels.showSerName = False
    #bar.dataLabels.showCatName = True  # Show attribute name at end
    #bar.dataLabels.position = 'r'  # Right end of bar
    bar.width = 34
    bar.height = 22
    # Add space between chart border and bars
    bar.gapWidth = 250
    bar.overlap = 0
    ws_chart = wb.create_sheet(chart_sheet_name)
    ws_chart.add_chart(bar, "B2")
    pct_ws.sheet_state = "hidden"
    wb.save(filepath)


